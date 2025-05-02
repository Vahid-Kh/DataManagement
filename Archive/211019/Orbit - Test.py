"""

This script is for processing the Excel files containing Hyperlinks and converting them into a .CSV extensions.
It may not be the most time efficient way of doing it but seems to be robust. (worked with all .xlsx file I thrown at it)

"""

import pandas as pd
import openpyxl
from funcPatent import df_patent_ipc

# from patent_scraping import add_abstract

inPath = r"C:\Users\U375297\Danfoss\RAC Tech Center - Programming Projects - Documents\Python Projects\Data management\Data\Raw Data\\"
wrdCldPath = r"C:\Users\U375297\Danfoss\RAC Tech Center - Programming Projects - Documents\Python Projects\Data management\Data\Word Cloud\\"
outPath = r"C:\Users\U375297\Danfoss\RAC Tech Center - Programming Projects - Documents\Python Projects\Data management\Data\Search Result\\"


# fileName ="Orbit - Data Sample - FD - Parker.xlsx"
# fileName ="Orbit - Data Sample - EEVs.xlsx"
# fileName ="Orbit - Test - EEVs.xlsx"
fileName ="Orbit - Test - EEVs - Copy.xlsx"
# fileName =".csv"
# fileName =".csv"
# fileName =".csv"
# fileName =".csv"
# fileName =".csv"
# fileName =".csv"
# fileName =".csv"
# fileName =".csv"
# fileName =".csv"


""" The column to read from """
# colName = "IPC Description"
# colName = "Title"
colName = "Abstract"
# colName = "Keywords"
# colName = "FullText"
"""Number of digits from IPC code"""
#N=3     #first level
#N=4     #second level
N=9      #complete IPC code

"""----------------------------------------------"""
""" SELECT SEARCH WORD :::   %LOWER CASE ONLY % """
""" Can only take lower case """
"""----------------------------------------------"""

""" TO PRINT FULL DATAFRAME   """
pd.set_option("display.max_rows", None, "display.max_columns", None)
pd.options.display.max_colwidth = 100
""" Iterate through the csv file  """


""" To load the workbook and the first sheet , change from 0 if needed"""
wb = openpyxl.load_workbook(inPath + fileName)
ws = wb[wb.sheetnames[0]]


print("File ", fileName, " has been loaded. ")
# print("Column name: ", colName)

"""  To get the tpe of data values and extract the hyperlinks"""

# print(ws.cell(row=2, column=4))
# print(ws.cell(row=2,column=4).value)
# print(ws.cell(row=2, column=5).hyperlink.target)
print("Dimension of the file loaded is :  ", ws.max_row -1," Rows", " * ",ws.max_column," Columns")

""" To get the dimension of te workbook """
row_num = ws.max_row
col_num = ws.max_column


rows = []
colval = []
colhyper =[]
cols = []

""" To create dataframe columns """
for i in range(col_num):
    try:
        trybait = ws.cell(row= 2, column=i + 1).hyperlink.target
        colhyper.append(ws.cell(row=1, column=i + 1).value)
    except:
        trybait = ws.cell(row=2, column=i + 1).value
        colval.append(ws.cell(row=1, column=i + 1).value)


cols.append(colhyper + colval)
df = pd.DataFrame(columns=cols[0])

""" to create a list of rows lists ( It rhymed a bit :) )"""
for j in range(row_num-1):
    hyper = []
    val = []
    for i in range(col_num):
        try:
            hyper.append(ws.cell(row = j+2, column=i + 1).hyperlink.target)
        except:
            val.append(ws.cell(row = j+2, column=i + 1).value)

    rows.append(hyper+val)

print("Rows has been extracted.")

for k in range(len(rows)):
    df.loc[len(df)] = rows[k]


print("Here is a list of columns available in dataframe : ", cols[0])
print(df[ 'Abstract'])
print(df["Publication numbers"])
print(df['IPC - International classification'])
df.rename(columns = {'IPC - International classification':"IPCR Classifications"}, inplace = True)
print(df["IPCR Classifications"])
df["IPCR Classifications"]=df["IPCR Classifications"].str.replace("\n",';;')
df["IPCR Classifications"]=df["IPCR Classifications"].str.replace("-0",'')
# df["IPCR Classifications"]=df["IPCR Classifications"].str.replace("/",'')
print(df["IPCR Classifications"])

df = df_patent_ipc(df,1)
print(df['IPC Description'])
df = df_patent_ipc(df,3)
print(df['IPC Description'])
df = df_patent_ipc(df,4)
print(df['IPC Description'])
df = df_patent_ipc(df,7)
print(df['IPC Description'])
df = df_patent_ipc(df,9)
print(df['IPC Description'])
